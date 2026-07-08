"""Phase 2 completion tests: blindfolding Q², NFI/RMS_theta, Excel export,
PDF gating, assumption gates, and the research-chat endpoint.

Blindfolding Q² is validated against the primer 2nd-edition corp-rep values
(CUSA 0.280, CUSL 0.415, omission distance 7).
"""
import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from backend.app import main as main_module
from backend.app.main import app
from backend.app.storage import ROOT

from .helpers import create_analysis

CSV = Path(__file__).parent / "fixtures_corp_rep.csv"

client = TestClient(app)


@pytest.fixture(scope="module")
def dataset():
    with CSV.open("rb") as f:
        resp = client.post(
            "/api/datasets",
            files={"file": ("corp_rep.csv", f, "text/csv")},
            data={"missing_value": "-99"},
        )
    assert resp.status_code == 200, resp.text
    return resp.json()


@pytest.fixture(scope="module")
def analysis(dataset):
    spec = json.loads((ROOT / "ai" / "fixtures" / "model_spec_reference.json").read_text())
    analysis_id = create_analysis(client, {
        "dataset_id": dataset["id"], "nboot": 100, "prediction": False,
        "constructs": [{"name": c["name"], "indicators": c.get("indicators", []),
                        "measurement": c["measurement"]} for c in spec["constructs"]],
        "paths": [{"from_construct": p["from_construct"], "to_construct": p["to_construct"]}
                  for p in spec["paths"]],
    })["id"]
    return {
        "id": analysis_id,
        "results": client.get(f"/api/analyses/{analysis_id}/results").json(),
        "assessment": client.get(f"/api/analyses/{analysis_id}/assessment").json(),
    }


def test_blindfolding_q2_matches_published(analysis):
    q2 = {r["row"]: r["q2"] for r in analysis["results"]["blindfolding"]}
    assert q2["CUSA"] == pytest.approx(0.280, abs=0.02)
    assert q2["CUSL"] == pytest.approx(0.415, abs=0.02)
    fam = {}
    for s in analysis["assessment"]["structural_model"]:
        fam.setdefault(s["family"], []).append(s)
    q2_verdicts = {s["construct"]: s["verdict"] for s in fam["predictive_relevance_q2"]}
    assert q2_verdicts["CUSL"] == "substantial"


def test_fit_indices(analysis):
    r = analysis["results"]
    assert 0 < r["nfi"] < 1
    assert 0 < r["rms_theta"] < 1
    fits = {s["metric"]: s for s in analysis["assessment"]["structural_model"]
            if s["family"] == "model_fit"}
    assert {"SRMR (saturated)", "NFI", "RMS_theta"} <= set(fits)
    assert all(f["verdict"] in ("pass", "review", "fail") for f in fits.values())


def test_full_collinearity_cmb(analysis):
    # Kock (2015) full-collinearity test: engine emits a VIF per construct and the
    # assessment turns each into a common_method_bias verdict (threshold 3.3).
    fc = analysis["results"]["full_collinearity_vif"]
    assert fc and all("construct" in r and "vif" in r for r in fc)
    cmb = [s for s in analysis["assessment"]["structural_model"]
           if s["family"] == "common_method_bias"]
    assert len(cmb) == len(fc)
    # corp-rep QUAL is collinear with the other drivers (VIF > 3.3) -> a fail.
    qual = next(s for s in cmb if s["construct"] == "QUAL")
    assert qual["value"] > 3.3 and qual["verdict"] == "fail"
    assert all(s["verdict"] in ("pass", "fail") for s in cmb)


def test_excel_export(analysis):
    resp = client.get(f"/api/analyses/{analysis['id']}/results.xlsx")
    assert resp.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    assert {"Model fit", "Paths and R2", "Hypotheses", "Blindfolding Q2",
            "Specific indirect", "IPMA performance"} <= set(wb.sheetnames)
    fit = {row[0].value: row[1].value for row in wb["Model fit"].iter_rows(min_row=2)}
    assert fit["SRMR"] == pytest.approx(analysis["results"]["srmr"], abs=1e-6)


def test_pdf_gated_on_libreoffice(analysis, monkeypatch):
    monkeypatch.setattr(main_module, "_find_soffice", lambda: None)
    resp = client.get(f"/api/analyses/{analysis['id']}/report.pdf")
    assert resp.status_code == 501
    assert "LibreOffice" in resp.json()["detail"]


def test_assumption_gates(dataset):
    # 8 formative indicators demand n >= 80 (10-times rule); serve 40 rows.
    import pandas as pd
    small = pd.read_csv(CSV).head(40)
    buf = io.BytesIO()
    small.to_csv(buf, index=False)
    buf.seek(0)
    resp = client.post("/api/datasets", files={"file": ("small.csv", buf, "text/csv")},
                       data={"missing_value": "-99"})
    small_ds = resp.json()

    spec = {
        "dataset_id": small_ds["id"], "nboot": 100, "prediction": False,
        "constructs": [
            {"name": "QUAL", "indicators": [f"qual_{i}" for i in range(1, 9)],
             "measurement": "formative"},
            {"name": "CUSA", "indicators": ["cusa"], "measurement": "single_item"},
        ],
        "paths": [{"from_construct": "QUAL", "to_construct": "CUSA"}],
    }
    resp = client.post("/api/analyses", json=spec)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["stage"] == "assumption_gates"
    assert any(v["gate"] == "sample_size_10x" for v in detail["violations"])

    meta = create_analysis(client, {**spec, "override_gates": True})
    assert meta["assumption_gates"]["overridden"] is True
    assert meta["assumption_gates"]["violations"]


def test_chat_gated_on_credentials(analysis, monkeypatch):
    from backend.app import ai
    monkeypatch.setattr(ai, "is_configured", lambda: False)
    resp = client.post(f"/api/analyses/{analysis['id']}/chat",
                       json={"message": "Why is HTMT below 0.85 good?"})
    assert resp.status_code == 503
    assert "credentials" in resp.json()["detail"]
